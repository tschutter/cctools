#!/usr/bin/env python2

"""
Generates a Purchase Order / Commercial Invoice.

All known products are included.  If you are never ordering a product
again, then just don't order it.  Or use the --exclude-sku option.
"""

from __future__ import print_function
import ConfigParser
import argparse
import cctools
import datetime
import itertools
import logging
import notify_send_handler
import openpyxl  # sudo pip install openpyxl
import os

CHECK_FOR_LACK_OF_ANY = False  # until most "Any" variants have been added

NUMBER_FORMAT_USD = "$#,##0.00;-$#,##0.00"

# Column numbers of product values.
COL_LINE_NO = 1
COL_SKU = 2
COL_DESCRIPTION = 3
COL_PRICE = 4
COL_QTY = 5
COL_TOTAL = 6
COL_HTSUS_NO = 7
COL_INSTRUCTIONS = 8


def set_cell(
    worksheet,
    row,
    col,
    value,
    font_bold=False,
    font_size=11,
    alignment_horizontal="general",
    alignment_vertical="bottom",
    number_format="General"
):
    """Set cell value and style."""
    cell = worksheet.cell(row=row, column=col)
    cell.value = value
    cell.font = openpyxl.styles.Font(bold=font_bold, size=font_size)
    cell.alignment = openpyxl.styles.Alignment(
        horizontal=alignment_horizontal,
        vertical=alignment_vertical
    )
    if number_format != "General":
        cell.number_format = number_format


def col_letter(col):
    """Return column letter for given column."""
    return chr(ord("A") + col - 1)


def add_title(worksheet):
    """Add worksheet title."""
    set_cell(worksheet, 1, 1, "Purchase Order", font_bold=True, font_size=20)
    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=3
    )
    worksheet.row_dimensions[1].height = 25


def set_label_value(
    worksheet,
    row,
    col_label_start,
    col_label_end,
    col_value_start,
    col_value_end,
    label,
    value
):
    """Add label: value."""
    if col_label_start != col_label_end:
        worksheet.merge_cells(
            start_row=row,
            start_column=col_label_start,
            end_row=row,
            end_column=col_label_end
        )
        col_label = col_label_start
    else:
        col_label = col_label_end
    set_cell(
        worksheet,
        row,
        col_label,
        label,
        font_bold=True,
        alignment_horizontal="right"
    )

    if col_value_start != col_value_end:
        worksheet.merge_cells(
            start_row=row,
            start_column=col_value_start,
            end_row=row,
            end_column=col_value_end
        )
        col_value = col_value_start
    else:
        col_value = col_value_end
    cell = worksheet.cell(row=row, column=col_value)
    cell.set_explicit_value(value)  # Force type to be string.


def add_header(args, config, worksheet, row):
    """Add PO/Invoice header."""

    col_label_start = 3
    col_label_end = 3
    col_value_start = col_label_end + 1
    col_value_end = col_value_start + 3

    set_label_value(
        worksheet,
        row - 1,
        col_label_start,
        col_label_end,
        col_value_start,
        col_value_end,
        "PO/Invoice #: ",
        args.number
    )
    row += 1

    date_str = datetime.date.today().strftime("%B %d, %Y")
    set_label_value(
        worksheet,
        row - 1,
        col_label_start,
        col_label_end,
        col_value_start,
        col_value_end,
        "Date: ",
        date_str
    )
    row += 1

    set_label_value(
        worksheet,
        row - 1,
        col_label_start,
        col_label_end,
        col_value_start,
        col_value_end,
        "Country of Origin: ",
        config.get("invoice", "country_of_origin")
    )
    row += 1

    set_label_value(
        worksheet,
        row - 1,
        col_label_start,
        col_label_end,
        col_value_start,
        col_value_end,
        "Manufacturer ID: ",
        config.get("invoice", "manufacturer_id")
    )
    row += 1

    first_row = True
    for key, value in config.items("invoice"):
        if key.startswith("consignee"):
            set_label_value(
                worksheet,
                row - 1,
                col_label_start,
                col_label_end,
                col_value_start,
                col_value_end,
                "Ultimate Consignee: " if first_row else "",
                value
            )
            first_row = False
            row += 1

    # Unit of measurement.
    set_label_value(
        worksheet,
        row - 1,
        col_label_start,
        col_label_end,
        col_value_start,
        col_value_end,
        "Unit of Measurement: ",
        config.get("invoice", "unit_of_measurement")
    )
    row += 1

    # Terms of sale.
    set_label_value(
        worksheet,
        row - 1,
        col_label_start,
        col_label_end,
        col_value_start,
        col_value_end,
        "Currency: ",
        config.get("invoice", "currency")
    )
    row += 1

    # Transport and delivery.
    set_label_value(
        worksheet,
        row - 1,
        col_label_start,
        col_label_end,
        col_value_start,
        col_value_end,
        "Transport and Delivery: ",
        config.get("invoice", "transport_and_delivery")
    )
    row += 1

    # Terms of sale.
    set_label_value(
        worksheet,
        row - 1,
        col_label_start,
        col_label_end,
        col_value_start,
        col_value_end,
        "Terms of Sale: ",
        config.get("invoice", "terms_of_sale")
    )
    row += 1

    return row


def add_variant(
    worksheet,
    row,
    lineno,
    sku,
    description,
    cost,
    htsus_no
):
    """Add a row for a variant."""

    set_cell(worksheet, row, COL_LINE_NO, lineno)
    set_cell(
        worksheet,
        row,
        COL_SKU,
        sku,
        alignment_horizontal="left"
    )
    set_cell(worksheet, row, COL_DESCRIPTION, description)
    set_cell(worksheet, row, COL_PRICE, cost, number_format=NUMBER_FORMAT_USD)
    total_formula = "=IF({}{}=\"\", \"\", {}{} * {}{})".format(
        col_letter(COL_QTY),
        row,
        col_letter(COL_PRICE),
        row,
        col_letter(COL_QTY),
        row
    )
    set_cell(
        worksheet,
        row,
        COL_TOTAL,
        total_formula,
        number_format=NUMBER_FORMAT_USD
    )
    if htsus_no is not None:
        set_cell(worksheet, row, COL_HTSUS_NO, htsus_no)


def get_product_variants(variants, sku):
    """Returns a list of variants for a product."""
    product_variants = [
        variant for variant in variants if variant["Product SKU"] == sku
    ]
    product_variants.sort(key=lambda variant: variant["Answer Sort Order"])
    return product_variants


def add_product(worksheet, row, lineno, product, variants):
    """Add row for each variant."""
    sku = product["SKU"]
    teaser = cctools.html_to_plain_text(product["Teaser"])
    cost = float(product["Cost"])
    if "HTSUS No" in product:
        htsus_no = product["HTSUS No"]
    else:
        htsus_no = None
    product_variants = get_product_variants(variants, sku)
    if len(product_variants) == 0:
        description = "{}: {}".format(
            product["Product Name"],
            teaser
        )
        add_variant(worksheet, row, lineno, sku, description, cost, htsus_no)
        row += 1
        lineno += 1
    else:
        any_variant_exists = False
        for variant in product_variants:
            if variant["SKU"] == "ANY" or variant["SKU"] == "VAR":
                any_variant_exists = True
            variant_sku = "{}-{}".format(sku, variant["SKU"])
            answer = variant["Question|Answer"].split("|")[1]
            description = "{} ({}): {}".format(
                product["Product Name"],
                answer,
                teaser
            )
            add_variant(
                worksheet,
                row,
                lineno,
                variant_sku,
                description,
                cost,
                htsus_no
            )
            row += 1
            lineno += 1
        if CHECK_FOR_LACK_OF_ANY and not any_variant_exists:
            logging.getLogger().warning(
                "No 'Any' or 'Variety' variant exists for {} {}".format(
                    sku,
                    product["Product Name"]
                )
            )

    return row, lineno


def add_products(worksheet, row, cc_browser, products, has_htsus_no):
    """Add row for each product."""

    # Add header row.
    set_cell(
        worksheet,
        row,
        COL_LINE_NO,
        "Line No",
        font_bold=True,
        alignment_horizontal="right"
    )
    set_cell(
        worksheet,
        row,
        COL_SKU,
        "SKU",
        font_bold=True,
        alignment_horizontal="left"
    )
    set_cell(worksheet, row, COL_DESCRIPTION, "Description", font_bold=True)
    set_cell(
        worksheet,
        row,
        COL_PRICE,
        "Price",
        font_bold=True,
        alignment_horizontal="right"
    )
    set_cell(
        worksheet,
        row,
        COL_QTY,
        "Qty",
        font_bold=True,
        alignment_horizontal="right"
    )
    set_cell(
        worksheet,
        row,
        COL_TOTAL,
        "Total",
        font_bold=True,
        alignment_horizontal="right"
    )
    if has_htsus_no:
        set_cell(
            worksheet,
            row,
            COL_HTSUS_NO,
            "HTSUS No",
            font_bold=True
        )
    set_cell(
        worksheet,
        row,
        COL_INSTRUCTIONS,
        "Item Special Instructions",
        font_bold=True
    )
    row += 1

    # Fetch variants list.
    variants = cc_browser.get_variants()

    # Group products by category.
    first_product_row = row
    lineno = 1
    for _, product_group in itertools.groupby(
        products,
        key=cc_browser.product_key_by_category
    ):
        # Leave a row for the category name.
        category = "unknown"
        category_row = row
        row += 1
        # Add product rows.
        for product in product_group:
            row, lineno = add_product(
                worksheet,
                row,
                lineno,
                product,
                variants
            )
            category = product["Category"]
        # Go back and insert category name.
        set_cell(
            worksheet,
            category_row,
            COL_DESCRIPTION,
            category,
            font_bold=True
        )
    last_product_row = row - 1

    # Set column widths.
    worksheet.column_dimensions[col_letter(COL_LINE_NO)].width = 8
    worksheet.column_dimensions[col_letter(COL_SKU)].width = 15
    worksheet.column_dimensions[col_letter(COL_DESCRIPTION)].width = 100
    worksheet.column_dimensions[col_letter(COL_PRICE)].width = 7
    worksheet.column_dimensions[col_letter(COL_QTY)].width = 5
    worksheet.column_dimensions[col_letter(COL_TOTAL)].width = 10
    if has_htsus_no:
        worksheet.column_dimensions[col_letter(COL_HTSUS_NO)].width = 13
    worksheet.column_dimensions[col_letter(COL_INSTRUCTIONS)].width = 30

    return row, first_product_row, last_product_row


def set_label_dollar_value(
    worksheet,
    row,
    col_label_start,
    col_label_end,
    col_total,
    label,
    value,
    font_bold=False
):
    """Add label: value."""
    worksheet.merge_cells(
        start_row=row,
        start_column=col_label_start,
        end_row=row,
        end_column=col_label_end
    )
    set_cell(
        worksheet,
        row,
        col_label_start,
        label,
        font_bold=font_bold,
        alignment_horizontal="right"
    )
    set_cell(
        worksheet,
        row,
        col_total,
        value,
        number_format=NUMBER_FORMAT_USD,
        font_bold=font_bold
    )


def add_totals(
    worksheet,
    config,
    row,
    col_total,
    first_product_row,
    last_product_row
):
    """Add subtotals and totals."""

    col_label_start = col_total - 3
    col_label_end = col_total - 1

    # Subtotal.
    subtotal_formula = "=SUM({}{}:{}{})".format(
        col_letter(col_total),
        first_product_row,
        col_letter(col_total),
        last_product_row
    )
    set_label_dollar_value(
        worksheet,
        row,
        col_label_start,
        col_label_end,
        col_total,
        "Subtotal:",
        subtotal_formula
    )
    subtotal_row = row
    last_adjustment_row = row
    row += 1

    # Discount.
    if config.has_option("invoice", "percent_discount"):
        percent_discount = config.getfloat("invoice", "percent_discount")
        discount_formula = "={}{} * {}".format(
            col_letter(col_total),
            subtotal_row,
            -percent_discount / 100.0
        )
        set_label_dollar_value(
            worksheet,
            row,
            col_label_start,
            col_label_end,
            col_total,
            "{}% Discount:".format(percent_discount),
            discount_formula
        )
        row += 1

    # Shipping.
    set_label_dollar_value(
        worksheet,
        row,
        col_label_start,
        col_label_end,
        col_total,
        "Shipping:",
        0.0
    )
    row += 1

    # Adjustments.
    last_adjustment_row = row - 1
    for key, value in config.items("invoice"):
        if key.startswith("adjustment"):
            set_label_dollar_value(
                worksheet,
                row,
                col_label_start,
                col_label_end,
                col_total,
                value + ":",
                ""
            )
            last_adjustment_row = row
            row += 1

    # Total.
    total_formula = "=SUM({}{}:{}{})".format(
        col_letter(col_total),
        subtotal_row,
        col_letter(col_total),
        last_adjustment_row
    )
    set_label_dollar_value(
        worksheet,
        row,
        col_label_start,
        col_label_end,
        col_total,
        "Total:",
        total_formula
    )
    row += 1

    return row


def add_summary(
    worksheet,
    row,
    first_product_row,
    last_product_row,
    htsus_numbers
):
    """
    Add a section that summarizes total cost by HTSUS number.  This
    is intended to make the customs inspector's job easier and
    therefore less likely to be grumpy and cause problems.
    """

    # Section title.
    set_cell(
        worksheet,
        row,
        COL_DESCRIPTION,
        "Summary by HTSUS number:",
        font_bold=True
    )
    row += 1

    for htsus_no in sorted(htsus_numbers):
        total_range = "{0}{1}:{0}{2}".format(
            col_letter(COL_TOTAL),
            first_product_row,
            last_product_row
        )
        htsus_no_range = "{0}{1}:{0}{2}".format(
            col_letter(COL_HTSUS_NO),
            first_product_row,
            last_product_row
        )
        if htsus_no == "":
            htsus_no = "Not assigned"
            # SUMIF() with "" test does not work.
            total_formula = "=SUMPRODUCT({}, {}=\"\")".format(
                total_range,
                htsus_no_range
            )
        else:
            total_formula = "=SUMIF({}, \"{}\", {})".format(
                htsus_no_range,
                htsus_no,
                total_range,
            )
        set_cell(
            worksheet,
            row,
            COL_TOTAL,
            total_formula,
            number_format=NUMBER_FORMAT_USD
        )
        set_cell(worksheet, row, COL_HTSUS_NO, htsus_no)
        row += 1

    return row


def add_special_instructions(worksheet, row):
    """Add special instructions."""
    col_value_name = 3

    # Special instructions.
    set_cell(
        worksheet,
        row,
        col_value_name,
        "General Special Instructions:",
        font_bold=True
    )


def add_invoice(args, config, cc_browser, worksheet):
    """Create the PO-Invoice worksheet."""

    # Prepare worksheet.
    worksheet.title = "PO-Invoice"

    # Add title.
    add_title(worksheet)

    # Add header.
    row = 3
    row = add_header(args, config, worksheet, row)

    # Blank row.
    row += 1

    # Fetch products list.
    products = cc_browser.get_products()

    # Remove excluded SKUs.
    if args.exclude_skus:
        products = [
            x for x in products if str(x["SKU"]) not in args.exclude_skus
        ]

    # Sort products by category, product_name.
    products = sorted(products, key=cc_browser.product_key_by_cat_and_name)

    # Determine if products include HTSUS number.
    has_htsus_no = len(products) > 0 and "HTSUS No" in products[0]

    # Add products.
    row, first_product_row, last_product_row = add_products(
        worksheet,
        row,
        cc_browser,
        products,
        has_htsus_no
    )

    # Blank row.
    row += 1

    # Add subtotals and total.
    row = add_totals(
        worksheet,
        config,
        row,
        COL_TOTAL,
        first_product_row,
        last_product_row
    )

    # Blank row.
    row += 1

    if has_htsus_no:
        # Create list of unique HTSUS numbers.
        htsus_numbers = []
        for product in products:
            if product["HTSUS No"] not in htsus_numbers:
                htsus_numbers.append(product["HTSUS No"])

        # Add summary by HTSUS.
        row = add_summary(
            worksheet,
            row,
            first_product_row,
            last_product_row,
            htsus_numbers
        )

    # Add special instructions.
    add_special_instructions(worksheet, row)


def add_instructions(config, worksheet):
    """Create the Instructions worksheet."""

    # Prepare worksheet.
    worksheet.title = "Instructions"
    col_instruction = 1
    row = 1

    # Add instructions.
    for key, value in config.items("invoice"):
        if key.startswith("instruction"):
            if value == "[EMPTY]":
                row += 1
                continue
            bold = False
            if value.startswith("[BOLD]"):
                bold = True
                value = value[6:]
            set_cell(worksheet, row, col_instruction, value, font_bold=bold)
            row += 1

    # Set column widths.
    worksheet.column_dimensions[col_letter(col_instruction)].width = 120


def generate_xlsx(args, config, cc_browser):
    """Generate the XLS file."""

    # Construct a document.
    workbook = openpyxl.workbook.Workbook()

    # Create PO-Invoice worksheet.
    add_invoice(args, config, cc_browser, workbook.worksheets[0])

    # Create Instructions worksheet.
    add_instructions(config, workbook.create_sheet())

    # Write to file.
    workbook.save(args.xlsx_filename)


def main():
    """main"""
    default_config = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "cctools.cfg"
    )
    now = datetime.datetime.now()
    default_number = now.strftime("%y%m%d00")
    default_xlsx_filename = now.strftime("%Y-%m-%d-PurchaseOrder.xlsx")

    arg_parser = argparse.ArgumentParser(
        description="Generates a Purchase Order / Commercial Invoice."
    )
    arg_parser.add_argument(
        "--config",
        action="store",
        dest="config",
        metavar="FILE",
        default=default_config,
        help="configuration filename (default=%(default)s)"
    )
    arg_parser.add_argument(
        "--number",
        action="store",
        dest="number",
        metavar="NUM",
        default=default_number,
        help="PO/Invoice number (default=%(default)s)"
    )
    arg_parser.add_argument(
        "--outfile",
        action="store",
        dest="xlsx_filename",
        metavar="FILE",
        default=default_xlsx_filename,
        help="output XLSX filename (default=%(default)s)"
    )
    arg_parser.add_argument(
        "--exclude-sku",
        action="append",
        dest="exclude_skus",
        metavar="SKU",
        help="exclude SKU from output"
    )
    arg_parser.add_argument(
        "--verbose",
        action="store_true",
        default=False,
        help="display progress messages"
    )

    # Parse command line arguments.
    args = arg_parser.parse_args()

    # Configure logging.
    logging.basicConfig(
        level=logging.INFO if args.verbose else logging.WARNING
    )
    logger = logging.getLogger()

    # Also log using notify-send if it is available.
    if notify_send_handler.NotifySendHandler.is_available():
        logger.addHandler(
            notify_send_handler.NotifySendHandler(
                os.path.splitext(os.path.basename(__file__))[0]
            )
        )

    # Read config file.
    config = ConfigParser.RawConfigParser()
    config.readfp(open(args.config))

    # Create a connection to CoreCommerce.
    cc_browser = cctools.CCBrowser(
        config.get("website", "host"),
        config.get("website", "site"),
        config.get("website", "username"),
        config.get("website", "password")
    )

    # Generate spreadsheet.
    logger.debug("Generating {}\n".format(os.path.abspath(args.xlsx_filename)))
    generate_xlsx(args, config, cc_browser)

    logger.debug("Generation complete")
    return 0


if __name__ == "__main__":
    main()
