#!/usr/bin/env python2

"""
Generates a FedEx Product Dictionary.
"""

import ConfigParser
import argparse
import cctools
import itertools
import logging
import notify_send_handler
import openpyxl  # sudo pip install openpyxl
import os
import datetime


def add_product_dict(args, cc_browser, products, worksheet):
    """Create the Product Dictionary worksheet."""

    # Prepare worksheet.
    worksheet.title = "Product Dictionary"

    # Add products.
    # Remove excluded SKUs.
    if args.exclude_skus:
        products = [
            x for x in products if str(x["SKU"]) not in args.exclude_skus
        ]

    # Add product rows, grouped by category.
    row = 1
    for _, product_group in itertools.groupby(
        products,
        key=cc_browser.product_key_by_category
    ):
        # Add product rows.
        for product in product_group:
            if product["Discontinued Item"] == "Y":
                continue
            description = "{}: {}".format(
                product["Product Name"],
                cctools.html_to_plain_text(product["Teaser"])
            )
            worksheet.cell(row=row, column=1).value = product["SKU"]
            worksheet.cell(row=row, column=2).value = description
            worksheet.cell(row=row, column=3).value = product["HTSUS No"]
            row += 1

    # Set column widths.
    worksheet.column_dimensions["A"].width = 6
    worksheet.column_dimensions["B"].width = 95
    worksheet.column_dimensions["C"].width = 13


def generate_xlsx(args, cc_browser, products):
    """Generate the XLS file."""

    # Construct a document.
    workbook = openpyxl.workbook.Workbook()

    # Create Product Dictionary worksheet.
    add_product_dict(args, cc_browser, products, workbook.worksheets[0])

    # Write to file.
    workbook.save(args.xlsx_filename)


def main():
    """main"""
    default_config = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "cctools.cfg"
    )
    now = datetime.datetime.now()
    default_xlsx_filename = now.strftime("%Y-%m-%d-PurchaseOrder.xlsx")

    arg_parser = argparse.ArgumentParser(
        description="Generates a FedEx Product Dictionary."
    )
    arg_parser.add_argument(
        "--config",
        action="store",
        dest="config",
        metavar="FILE",
        default=default_config,
        help="configuration filename (default=%(default)s)"
    )
    arg_parser.add_argument(
        "--outfile",
        action="store",
        dest="xlsx_filename",
        metavar="FILE",
        default=default_xlsx_filename,
        help="output XLSX filename (default=%(default)s)"
    )
    arg_parser.add_argument(
        "--exclude-sku",
        action="append",
        dest="exclude_skus",
        metavar="SKU",
        help="exclude SKU from output"
    )
    arg_parser.add_argument(
        "--verbose",
        action="store_true",
        default=False,
        help="display progress messages"
    )

    # Parse command line arguments.
    args = arg_parser.parse_args()

    # Configure logging.
    logging.basicConfig(
        level=logging.INFO if args.verbose else logging.WARNING
    )
    logger = logging.getLogger()

    # Also log using notify-send if it is available.
    if notify_send_handler.NotifySendHandler.is_available():
        logger.addHandler(
            notify_send_handler.NotifySendHandler(
                os.path.splitext(os.path.basename(__file__))[0]
            )
        )

    # Read config file.
    config = ConfigParser.RawConfigParser()
    config.readfp(open(args.config))

    # Create a connection to CoreCommerce.
    cc_browser = cctools.CCBrowser(
        config.get("website", "base_url"),
        config.get("website", "username"),
        config.get("website", "password")
    )

    # Fetch products list.
    products = cc_browser.get_products()

    # Generate spreadsheet.
    logger.debug("Generating {}".format(os.path.abspath(args.xlsx_filename)))
    generate_xlsx(args, cc_browser, products)

    logger.debug("Generation complete")
    return 0


if __name__ == "__main__":
    main()
