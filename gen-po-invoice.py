#!/usr/bin/env python2

"""
Generates a Purchase Order / Commercial Invoice.
"""

import ConfigParser
import argparse
import cctools
import itertools
import openpyxl  # sudo apt-get install python-openpyxl
import os
import sys
import datetime

# Cell style constants.
ALIGNMENT_HORIZONTAL_RIGHT = openpyxl.style.Alignment.HORIZONTAL_RIGHT
ALIGNMENT_VERTICAL_TOP = openpyxl.style.Alignment.VERTICAL_TOP
NUMBER_FORMAT_USD = openpyxl.style.NumberFormat.FORMAT_CURRENCY_USD_SIMPLE

def has_merge_cells(worksheet):
    """Determine if Worksheet.merge_cells method exists."""
    # merge_cells not supported by openpyxl-1.5.6 (Ubuntu 12.04)
    return hasattr(worksheet, "merge_cells")

def set_cell(
    worksheet,
    row,
    col,
    value,
    bold = None,
    alignment_horizontal = None,
    alignment_vertical = None
):
    """Set cell value and style."""
    cell = worksheet.cell(row=row, column=col)
    cell.value = value
    if bold != None:
        cell.style.font.bold = bold
    if alignment_horizontal != None:
        cell.style.alignment.horizontal = alignment_horizontal
    if alignment_vertical != None:
        cell.style.alignment.vertical = alignment_vertical
    return cell


def col_letter(col):
    """Return column letter for given column."""
    return chr(ord("A") + col)


def row_number(row):
    """Return row number for given row."""
    return row + 1


def add_title(worksheet):
    """Add worksheet title."""
    style = set_cell(worksheet, 0, 0, "Purchase Order", bold=True).style
    style.font.size = 20
    if has_merge_cells(worksheet):
        worksheet.merge_cells(
            start_row=0,
            start_column=0,
            end_row=0,
            end_column=2
        )
    worksheet.row_dimensions[1].height = 25


def set_label_value(
    worksheet,
    row,
    col_label_start,
    col_label_end,
    col_value_start,
    col_value_end,
    label,
    value
):
    """Add label: value."""
    if has_merge_cells(worksheet) and col_label_start != col_label_end:
        worksheet.merge_cells(
            start_row=row,
            start_column=col_label_start,
            end_row=row,
            end_column=col_label_end
        )
        col_label = col_label_start
    else:
        col_label = col_label_end
    label_style = set_cell(worksheet, row, col_label, label, bold=True).style
    label_style.alignment.horizontal = ALIGNMENT_HORIZONTAL_RIGHT

    if has_merge_cells(worksheet) and col_value_start != col_value_end:
        worksheet.merge_cells(
            start_row=row,
            start_column=col_value_start,
            end_row=row,
            end_column=col_value_end
        )
        col_value = col_value_start
    else:
        col_value = col_value_end
    cell = worksheet.cell(row=row, column=col_value)
    cell.set_value_explicit(value)  # Force type to be string.


def add_header(args, config, worksheet, row):
    """Add PO/Invoice header."""
    col_value_name = 2
    col_value = col_value_name + 1

    col_label_start = 2
    col_label_end = 2
    col_value_start = col_label_end + 1
    col_value_end = col_value_start + 3

    set_label_value(
        worksheet,
        row - 1,
        col_label_start,
        col_label_end,
        col_value_start,
        col_value_end,
        "PO/Invoice #: ",
        args.number
    )
    row += 1

    date_str = datetime.date.today().strftime("%B %d, %Y")
    set_label_value(
        worksheet,
        row - 1,
        col_label_start,
        col_label_end,
        col_value_start,
        col_value_end,
        "Date: ",
        date_str
    )
    row += 1

    set_label_value(
        worksheet,
        row - 1,
        col_label_start,
        col_label_end,
        col_value_start,
        col_value_end,
        "Country of Origin: ",
        config.get("invoice", "country_of_origin")
    )
    row += 1

    set_label_value(
        worksheet,
        row - 1,
        col_label_start,
        col_label_end,
        col_value_start,
        col_value_end,
        "Manufacturer ID: ",
        config.get("invoice", "manufacturer_id")
    )
    row += 1

    first_row = True
    for key, value in config.items("invoice"):
        if key.startswith("consignee"):
            set_label_value(
                worksheet,
                row - 1,
                col_label_start,
                col_label_end,
                col_value_start,
                col_value_end,
                "Ultimate Consignee: " if first_row else "",
                value
            )
            first_row = False
            row += 1

    # Unit of measurement.
    set_label_value(
        worksheet,
        row - 1,
        col_label_start,
        col_label_end,
        col_value_start,
        col_value_end,
        "Unit of Measurement: ",
        config.get("invoice", "unit_of_measurement")
    )
    row += 1

    # Terms of sale.
    set_label_value(
        worksheet,
        row - 1,
        col_label_start,
        col_label_end,
        col_value_start,
        col_value_end,
        "Currency: ",
        config.get("invoice", "currency")
    )
    row += 1

    # Transport and delivery.
    set_label_value(
        worksheet,
        row - 1,
        col_label_start,
        col_label_end,
        col_value_start,
        col_value_end,
        "Transport and Delivery: ",
        config.get("invoice", "transport_and_delivery")
    )
    row += 1

    # Terms of sale.
    set_label_value(
        worksheet,
        row - 1,
        col_label_start,
        col_label_end,
        col_value_start,
        col_value_end,
        "Terms of Sale: ",
        config.get("invoice", "terms_of_sale")
    )
    row += 1

    return row


def add_products(args, worksheet, row, cc_browser, products):
    """Add row for each product."""
    col_line_no = 0
    col_sku = 1
    col_description = 2
    col_price = 3
    col_qty = 4
    col_total = 5
    col_htsus_no = 6
    col_instructions = 7

    # Add header row.
    set_cell(
        worksheet,
        row,
        col_line_no,
        "Line No",
        bold=True,
        alignment_horizontal=ALIGNMENT_HORIZONTAL_RIGHT
    )
    set_cell(
        worksheet,
        row,
        col_sku,
        "SKU",
        bold=True,
        alignment_horizontal=ALIGNMENT_HORIZONTAL_RIGHT
    )
    set_cell(worksheet, row, col_description, "Description", bold=True)
    set_cell(
        worksheet,
        row,
        col_price,
        "Price",
        bold=True,
        alignment_horizontal=ALIGNMENT_HORIZONTAL_RIGHT
    )
    set_cell(
        worksheet,
        row,
        col_qty,
        "Qty",
        bold=True,
        alignment_horizontal=ALIGNMENT_HORIZONTAL_RIGHT
    )
    set_cell(
        worksheet,
        row,
        col_total,
        "Total",
        bold=True,
        alignment_horizontal=ALIGNMENT_HORIZONTAL_RIGHT
    )
    set_cell(
        worksheet,
        row,
        col_htsus_no,
        "HTSUS No",
        bold=True
    )
    set_cell(
        worksheet,
        row,
        col_instructions,
        "Item Special Instructions",
        bold=True
    )
    row += 1

    # Remove excluded SKUs.
    if args.exclude_skus:
        products = [x for x in products if str(x["SKU"]) not in args.exclude_skus]

    # Sort products by category, product_name.
    products = sorted(products, key=cc_browser.sort_key_by_category_and_name)

    # Group products by category.
    first_product_row = row
    lineno = 1
    for _, product_group in itertools.groupby(
        products,
        key=cc_browser.sort_key_by_category
    ):
        # Leave a row for the category name.
        category = "unknown"
        category_row = row
        row += 1
        # Add product rows.
        for product in product_group:
            category = product["Category"]
            if product["Discontinued Item"] == "Y":
                continue
            description = "  %s: %s" % (
                product["Product Name"],
                cctools.html_to_plain_text(product["Teaser"])
            )
            htsus_no = product["HTSUS No"]
            set_cell(worksheet, row, col_line_no, lineno)
            set_cell(worksheet, row, col_sku, product["SKU"])
            set_cell(worksheet, row, col_description, description)
            style = set_cell(worksheet, row, col_price, product["Cost"]).style
            style.number_format.format_code = NUMBER_FORMAT_USD
            total_formula = "=IF(%s%i=\"\", \"\", %s%i * %s%i)" % (
                col_letter(col_qty),
                row_number(row),
                col_letter(col_price),
                row_number(row),
                col_letter(col_qty),
                row_number(row)
            )
            style = set_cell(worksheet, row, col_total, total_formula).style
            style.number_format.format_code = NUMBER_FORMAT_USD
            set_cell(worksheet, row, col_htsus_no, htsus_no)
            row += 1
            lineno += 1
        # Go back and insert category name.
        set_cell(worksheet, category_row, col_description, category, bold=True)
    last_product_row = row - 1

    # Set column widths.
    worksheet.column_dimensions[col_letter(col_line_no)].width = 8
    worksheet.column_dimensions[col_letter(col_sku)].width = 6
    worksheet.column_dimensions[col_letter(col_description)].width = 66
    worksheet.column_dimensions[col_letter(col_price)].width = 7
    worksheet.column_dimensions[col_letter(col_qty)].width = 5
    worksheet.column_dimensions[col_letter(col_total)].width = 10
    worksheet.column_dimensions[col_letter(col_htsus_no)].width = 12
    worksheet.column_dimensions[col_letter(col_instructions)].width = 30

    return row, col_total, first_product_row, last_product_row


def set_label_dollar_value(
    worksheet,
    row,
    col_label_start,
    col_label_end,
    col_total,
    label,
    value
):
    """Add label: value."""
    if has_merge_cells(worksheet):
        worksheet.merge_cells(
            start_row=row,
            start_column=col_label_start,
            end_row=row,
            end_column=col_label_end
        )
        col_label = col_label_start
    else:
        col_label = col_label_end
    label_style = set_cell(worksheet, row, col_label, label).style
    label_style.alignment.horizontal = ALIGNMENT_HORIZONTAL_RIGHT

    value_style = set_cell(worksheet, row, col_total, value).style
    value_style.number_format.format_code = NUMBER_FORMAT_USD
    return(label_style, value_style)


def add_totals(
    worksheet,
    config,
    row,
    col_total,
    first_product_row,
    last_product_row
):
    """Add subtotals and totals."""

    col_label_start = col_total - 3
    col_label_end = col_total - 1

    # Subtotal.
    subtotal_formula = "=SUM(%s%i:%s%i)" % (
        col_letter(col_total),
        row_number(first_product_row),
        col_letter(col_total),
        row_number(last_product_row)
    )
    set_label_dollar_value(
        worksheet,
        row,
        col_label_start,
        col_label_end,
        col_total,
        "Subtotal:",
        subtotal_formula
    )
    subtotal_row = row
    row += 1

    # Discount.
    percent_discount = config.getfloat("invoice", "percent_discount")
    discount_formula = "=%s%i * %f" % (
        col_letter(col_total),
        row_number(subtotal_row),
        -percent_discount / 100.0
    )
    set_label_dollar_value(
        worksheet,
        row,
        col_label_start,
        col_label_end,
        col_total,
        "{}% Discount:".format(percent_discount),
        discount_formula
    )
    last_adjustment_row = row
    row += 1

    # Adjustments.
    for key, value in config.items("invoice"):
        if key.startswith("adjustment"):
            set_label_dollar_value(
                worksheet,
                row,
                col_label_start,
                col_label_end,
                col_total,
                value + ":",
                ""
            )
            last_adjustment_row = row
            row += 1

    # Total.
    total_formula = "=SUM(%s%i:%s%i)" % (
        col_letter(col_total),
        row_number(subtotal_row),
        col_letter(col_total),
        row_number(last_adjustment_row)
    )
    set_label_dollar_value(
        worksheet,
        row,
        col_label_start,
        col_label_end,
        col_total,
        "Total:",
        total_formula
    )
    row += 1

    return(row)


def add_special_instructions(worksheet, row):
    """Add special instructions."""
    col_value_name = 2

    # Special instructions.
    set_cell(
        worksheet,
        row,
        col_value_name,
        "General Special Instructions:",
        bold=True
    )


def add_invoice(args, config, cc_browser, products, worksheet):
    """Create the PO-Invoice worksheet."""

    # Prepare worksheet.
    worksheet.title = "PO-Invoice"

    # Add title.
    add_title(worksheet)

    # Add header.
    row = 2
    row = add_header(args, config, worksheet, row)

    # Blank row.
    row += 1

    # Add products.
    row, col_total, first_product_row, last_product_row = add_products(
        args,
        worksheet,
        row,
        cc_browser,
        products
    )

    # Blank row.
    row += 1

    # Add subtotals and total.
    row = add_totals(
        worksheet,
        config,
        row,
        col_total,
        first_product_row,
        last_product_row
    )

    # Add special instructions.
    add_special_instructions(worksheet, row_number(row))


def add_instructions(config, worksheet):
    """Create the Instructions worksheet."""

    # Prepare worksheet.
    worksheet.title = "Instructions"
    col_instruction = 0
    row = 0

    # Add instructions.
    for key, value in config.items("invoice"):
        if key.startswith("instruction"):
            if value == "[EMPTY]":
                row += 1
                continue
            bold = False
            if value.startswith("[BOLD]"):
                bold = True
                value = value[6:]
            cell = set_cell(worksheet, row, col_instruction, value)
            if bold:
                cell.style.font.bold = bold
            row += 1

    # Set column widths.
    worksheet.column_dimensions[col_letter(col_instruction)].width = 70


def generate_xlsx(args, config, cc_browser, products):
    """Generate the XLS file."""

    # Construct a document.
    workbook = openpyxl.workbook.Workbook()

    # Create PO-Invoice worksheet.
    add_invoice(args, config, cc_browser, products, workbook.worksheets[0])

    # Create Instructions worksheet.
    add_instructions(config, workbook.create_sheet())

    # Write to file.
    workbook.save(args.xlsx_filename)


def main():
    """main"""
    defaultConfig = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "cctools.cfg"
    )
    now = datetime.datetime.now()
    default_number = now.strftime("%y%m%d00")
    default_xlsx_filename = now.strftime("%Y-%m-%d-PurchaseOrder.xlsx")

    arg_parser = argparse.ArgumentParser(
        description="Generates a Purchase Order / Commercial Invoice."
    )
    arg_parser.add_argument(
        "--config",
        action="store",
        dest="config",
        metavar="FILE",
        default=defaultConfig,
        help="configuration filename (default=%(default)s)"
    )
    arg_parser.add_argument(
        "--number",
        action="store",
        dest="number",
        metavar="NUM",
        default=default_number,
        help="PO/Invoice number (default=%(default)s)"
    )
    arg_parser.add_argument(
        "--outfile",
        action="store",
        dest="xlsx_filename",
        metavar="FILE",
        default=default_xlsx_filename,
        help="output XLSX filename (default=%(default)s)"
    )
    arg_parser.add_argument(
        "--exclude-sku",
        action="append",
        dest="exclude_skus",
        metavar="SKU",
        help="exclude SKU from output"
    )
    arg_parser.add_argument(
        "--verbose",
        action="store_true",
        default=False,
        help="display progress messages"
    )

    # Parse command line arguments.
    args = arg_parser.parse_args()

    # Read config file.
    config = ConfigParser.RawConfigParser()
    config.readfp(open(args.config))

    # Create a connection to CoreCommerce.
    cc_browser = cctools.CCBrowser(
        config.get("website", "host"),
        config.get("website", "site"),
        config.get("website", "username"),
        config.get("website", "password"),
        verbose=args.verbose
    )

    # Fetch products list.
    products = cc_browser.get_products()

    # Generate spreadsheet.
    if args.verbose:
        sys.stderr.write(
            "Generating %s\n" % os.path.abspath(args.xlsx_filename)
        )
    generate_xlsx(args, config, cc_browser, products)

    if args.verbose:
        sys.stderr.write("Generation complete\n")
    return 0


if __name__ == "__main__":
    main()
