#!/usr/bin/env python2

"""
Generates a wholesale order form in spreadsheet form.
"""

import ConfigParser
import argparse
import datetime
import itertools
import logging
import math
import os

import openpyxl  # sudo pip install openpyxl

import cctools
import notify_send_handler

CHECK_FOR_LACK_OF_ANY = False  # until most "Any" variants have been added

NUMBER_FORMAT_USD = "$#,##0.00;-$#,##0.00"

# Column numbers of product values.
COL_ITEM_NO = 1
COL_DESCRIPTION = 2
COL_PRICE = 3
COL_QTY = 4
COL_TOTAL = 5
COL_SKU = 6
COL_SIZE = 7


def set_cell(
    worksheet,
    row,
    col,
    value,
    font_bold=False,
    font_size=11,
    alignment_horizontal="general",
    alignment_vertical="bottom",
    number_format="General"
):
    """Set cell value and style."""
    cell = worksheet.cell(row=row, column=col)
    cell.value = value
    cell.font = openpyxl.styles.Font(bold=font_bold, size=font_size)
    cell.alignment = openpyxl.styles.Alignment(
        horizontal=alignment_horizontal,
        vertical=alignment_vertical
    )
    if number_format != "General":
        cell.number_format = number_format


def col_letter(col):
    """Return column letter for given column."""
    return chr(ord("A") + col - 1)


def add_title(args, config, worksheet):
    """Add worksheet title."""
    row = 1
    doc_title = config.get("wholesale_order", "title")
    set_cell(worksheet, row, 1, doc_title, font_bold=True, font_size=20)
    worksheet.row_dimensions[1].height = 25
    row += 1

    now = datetime.datetime.now()
    cell_text = now.strftime("Date: %Y-%m-%d")
    set_cell(worksheet, row, 1, cell_text)
    row += 1

    valid_date = now + datetime.timedelta(days=args.valid_ndays)
    cell_text = valid_date.strftime("Valid until: %Y-%m-%d")
    set_cell(worksheet, row, 1, cell_text)
    row += 1

    cell_text = "Prices are {:.0%} of retail".format(args.wholesale_fraction)
    set_cell(worksheet, row, 1, cell_text)
    row += 1

    for merge_row in range(1, row):
        worksheet.merge_cells(
            start_row=merge_row,
            start_column=1,
            end_row=merge_row,
            end_column=2
        )

    return row


def add_ship_to(worksheet, row):
    """Add Ship To block."""
    start_col = 1
    end_col = 2
    worksheet.merge_cells(
        start_row=row,
        start_column=start_col,
        end_row=row,
        end_column=end_col
    )
    set_cell(
        worksheet,
        row,
        start_col,
        "Ship To:",
        font_bold=True,
        alignment_horizontal="left"
    )
    row += 1

    nrows = 3
    for i in range(0, nrows):
        cell = worksheet.cell(row=row, column=start_col)
        cell.alignment = openpyxl.styles.Alignment(horizontal="left")
        worksheet.merge_cells(
            start_row=row,
            start_column=start_col,
            end_row=row,
            end_column=end_col
        )
        for col in range(start_col, end_col + 1):
            default_side = openpyxl.styles.Side(
                border_style=None,
                color='FF000000'
            )
            if i == 0:
                top = openpyxl.styles.Side("thin")  # BORDER_THIN
            else:
                top = default_side
            if i == nrows - 1:
                bottom = openpyxl.styles.Side("thin")
            else:
                bottom = default_side
            if col == start_col:
                left = openpyxl.styles.Side("thin")
            else:
                left = default_side
            if col == end_col:
                right = openpyxl.styles.Side("thin")
            else:
                right = default_side
            cell = worksheet.cell(row=row, column=col)
            cell.border = openpyxl.styles.Border(
                left=left,
                right=right,
                top=top,
                bottom=bottom
            )
        row += 1

    return row


def set_label_dollar_value(
    worksheet,
    row,
    col_label_start,
    col_label_end,
    col_total,
    label,
    value,
    font_bold=False
):
    """Add label: value."""
    worksheet.merge_cells(
        start_row=row,
        start_column=col_label_start,
        end_row=row,
        end_column=col_label_end
    )
    set_cell(
        worksheet,
        row,
        col_label_start,
        label,
        font_bold=font_bold,
        alignment_horizontal="right"
    )

    set_cell(
        worksheet,
        row,
        col_total,
        value,
        font_bold=font_bold,
        number_format=NUMBER_FORMAT_USD
    )


def add_variant(
    worksheet,
    row,
    item_no,
    size,
    sku,
    description,
    wholesale_price
):
    """Add a row for a variant."""

    set_cell(worksheet, row, COL_ITEM_NO, item_no)
    set_cell(worksheet, row, COL_DESCRIPTION, description)
    set_cell(
        worksheet,
        row,
        COL_PRICE,
        wholesale_price,
        number_format=NUMBER_FORMAT_USD
    )
    total_formula = "=IF({}{}=\"\", \"\", {}{} * {}{})".format(
        col_letter(COL_QTY),
        row,
        col_letter(COL_PRICE),
        row,
        col_letter(COL_QTY),
        row
    )
    set_cell(
        worksheet,
        row,
        COL_TOTAL,
        total_formula,
        number_format=NUMBER_FORMAT_USD
    )
    set_cell(worksheet, row, COL_SIZE, size)
    set_cell(worksheet, row, COL_SKU, sku)


def get_product_variants(variants, sku):
    """Returns a list of variants for a product."""
    product_variants = [
        variant for variant in variants
        if variant["Product SKU"] == sku and variant["Variant Enabled"] == "Y"
    ]
    product_variants.sort(key=lambda variant: variant["Variant Sort"])
    return product_variants


def calc_wholesale_price(args, price):
    """Calculate wholesale price based on retail price."""
    if price > 1.0:
        rounded_price = math.floor(price + 0.5)
    else:
        rounded_price = price
    return rounded_price * args.wholesale_fraction


def add_product(args, worksheet, row, item_no, product, variants):
    """Add row for each variant."""
    size = product["Size"]
    product_name = product["Product Name"]
    sku = product["SKU"]
    teaser = cctools.html_to_plain_text(product["Teaser"])
    price = float(product["Price"])

    product_variants = get_product_variants(variants, sku)
    if len(product_variants) == 0:
        description = "{}: {}".format(product_name, teaser)
        add_variant(
            worksheet,
            row,
            item_no,
            size,
            sku,
            description,
            calc_wholesale_price(args, price)
        )
        row += 1
        item_no += 1
    else:
        any_variant_exists = False
        for variant in product_variants:
            variant_sku = variant["Variant SKU"]
            if variant_sku == "ANY" or variant_sku == "VAR":
                any_variant_exists = True
            variant_sku = "{}-{}".format(sku, variant_sku)
            variant_add_price = float(variant["Variant Add Price"])
            variant_name = variant["Variant Name"]
            description = "{} ({}): {}".format(
                product_name,
                variant_name,
                teaser
            )
            add_variant(
                worksheet,
                row,
                item_no,
                size,
                variant_sku,
                description,
                calc_wholesale_price(args, price + variant_add_price)
            )
            row += 1
            item_no += 1

        if CHECK_FOR_LACK_OF_ANY and not any_variant_exists:
            logging.getLogger().warning(
                "No 'Any' or 'Variety' variant exists for {} {}".format(
                    sku,
                    product_name
                )
            )

    return row, item_no


def add_products(args, worksheet, row, cc_browser, products):
    """Add row for each product."""

    # Add header row.
    set_cell(
        worksheet,
        row,
        COL_ITEM_NO,
        "Item No",
        font_bold=True,
        alignment_horizontal="right"
    )
    set_cell(worksheet, row, COL_DESCRIPTION, "Description", font_bold=True)
    set_cell(
        worksheet,
        row,
        COL_PRICE,
        "Price",
        font_bold=True,
        alignment_horizontal="right"
    )
    set_cell(
        worksheet,
        row,
        COL_QTY,
        "Qty",
        font_bold=True,
        alignment_horizontal="right"
    )
    set_cell(
        worksheet,
        row,
        COL_TOTAL,
        "Total",
        font_bold=True,
        alignment_horizontal="right"
    )
    set_cell(
        worksheet,
        row,
        COL_SKU,
        "SKU",
        font_bold=True,
        alignment_horizontal="right"
    )
    set_cell(
        worksheet,
        row,
        COL_SIZE,
        "Size",
        font_bold=True
    )
    row += 1

    # Remove excluded SKUs.
    if args.exclude_skus:
        products = [
            x for x in products if str(x["SKU"]) not in args.exclude_skus
        ]

    # Sort products by category, product_name.
    products = sorted(products, key=cc_browser.product_key_by_cat_and_name)

    # Fetch variants list.
    variants = cc_browser.get_variants()

    # Group products by category.
    first_product_row = row
    item_no = 1
    for _, product_group in itertools.groupby(
        products,
        key=cc_browser.product_key_by_category
    ):
        # Leave a row for the category name.
        category = "unknown"
        category_row = row
        row += 1

        # Add product rows.
        for product in product_group:
            if product["Available"] != "Y":
                continue

            row, item_no = add_product(
                args,
                worksheet,
                row,
                item_no,
                product,
                variants
            )
            category = product["Category"]
            last_product_row = row - 1

        # Go back and insert category name.
        if category == "":
            category = "Uncategorized"
        set_cell(
            worksheet,
            category_row,
            COL_DESCRIPTION,
            category,
            font_bold=True
        )

    # Set column widths.
    worksheet.column_dimensions[col_letter(COL_ITEM_NO)].width = 8
    worksheet.column_dimensions[col_letter(COL_DESCRIPTION)].width = 100
    worksheet.column_dimensions[col_letter(COL_PRICE)].width = 8
    worksheet.column_dimensions[col_letter(COL_QTY)].width = 5
    worksheet.column_dimensions[col_letter(COL_TOTAL)].width = 10
    worksheet.column_dimensions[col_letter(COL_SKU)].width = 14
    worksheet.column_dimensions[col_letter(COL_SIZE)].width = 28

    # Blank row.
    row += 1

    col_label_start = COL_TOTAL - 2
    col_label_end = COL_TOTAL - 1

    # Subtotal.
    subtotal_formula = "=SUM({}{}:{}{})".format(
        col_letter(COL_TOTAL),
        first_product_row,
        col_letter(COL_TOTAL),
        last_product_row
    )
    set_label_dollar_value(
        worksheet,
        row,
        col_label_start,
        col_label_end,
        COL_TOTAL,
        "Subtotal:",
        subtotal_formula
    )
    subtotal_row = row
    row += 1

    # Shipping.
    set_label_dollar_value(
        worksheet,
        row,
        col_label_start,
        col_label_end,
        COL_TOTAL,
        "Shipping:",
        0.0
    )
    row += 1

    # Adjustment.
    set_label_dollar_value(
        worksheet,
        row,
        col_label_start,
        col_label_end,
        COL_TOTAL,
        "Adjustment:",
        0.0
    )
    row += 1

    # Total.
    total_formula = "=SUM({}{}:{}{})".format(
        col_letter(COL_TOTAL),
        subtotal_row,
        col_letter(COL_TOTAL),
        row - 1
    )
    set_label_dollar_value(
        worksheet,
        row,
        col_label_start,
        col_label_end,
        COL_TOTAL,
        "Total:",
        total_formula,
        font_bold=True
    )


def add_order_form(args, config, cc_browser, products, worksheet):
    """Create the Wholesale Order Form worksheet."""

    # Prepare worksheet.
    worksheet.title = "Wholesale Order Form"

    # Add title.
    row = add_title(args, config, worksheet)

    # Blank row.
    row += 1

    # Ship To block.
    row = add_ship_to(worksheet, row)

    # Blank row.
    row += 1

    # Add products.
    add_products(
        args,
        worksheet,
        row,
        cc_browser,
        products
    )


def generate_xlsx(args, config, cc_browser, products):
    """Generate the XLS file."""

    # Construct a document.
    workbook = openpyxl.workbook.Workbook()

    # Create PO-Invoice worksheet.
    add_order_form(
        args,
        config,
        cc_browser,
        products,
        workbook.worksheets[0]
    )

    # Write to file.
    workbook.save(args.xlsx_filename)


def main():
    """main"""
    default_config = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "cctools.cfg"
    )
    now = datetime.datetime.now()
    default_xlsx_filename = now.strftime("%Y-%m-%d-WholesaleOrderForm.xlsx")

    arg_parser = argparse.ArgumentParser(
        description="Generates a wholesale order form."
    )
    arg_parser.add_argument(
        "--config",
        action="store",
        dest="config",
        metavar="FILE",
        default=default_config,
        help="configuration filename (default=%(default)s)"
    )
    arg_parser.add_argument(
        "--wholesale-fraction",
        metavar="FRAC",
        default=0.5,
        help="wholesale price fraction (default=%(default).2f)"
    )
    arg_parser.add_argument(
        "--valid-ndays",
        metavar="N",
        type=int,
        default=30,
        help="number of days prices are valid (default=%(default)i)"
    )
    arg_parser.add_argument(
        "--outfile",
        action="store",
        dest="xlsx_filename",
        metavar="FILE",
        default=default_xlsx_filename,
        help="output XLSX filename (default=%(default)s)"
    )
    arg_parser.add_argument(
        "--exclude-sku",
        action="append",
        dest="exclude_skus",
        metavar="SKU",
        help="exclude SKU from output"
    )
    arg_parser.add_argument(
        "--verbose",
        action="store_true",
        default=False,
        help="display progress messages"
    )

    # Parse command line arguments.
    args = arg_parser.parse_args()

    # Configure logging.
    logging.basicConfig(
        level=logging.INFO if args.verbose else logging.WARNING
    )
    logger = logging.getLogger()

    # Also log using notify-send if it is available.
    if notify_send_handler.NotifySendHandler.is_available():
        logger.addHandler(
            notify_send_handler.NotifySendHandler(
                os.path.splitext(os.path.basename(__file__))[0]
            )
        )

    # Read config file.
    config = ConfigParser.RawConfigParser()
    config.readfp(open(args.config))

    # Create a connection to CoreCommerce.
    cc_browser = cctools.CCBrowser(
        config.get("website", "host"),
        config.get("website", "site"),
        config.get("website", "username"),
        config.get("website", "password")
    )

    # Fetch products list.
    products = cc_browser.get_products()

    # Generate spreadsheet.
    logger.debug("Generating {}".format(args.xlsx_filename))
    generate_xlsx(args, config, cc_browser, products)

    logger.debug("Generation complete")
    return 0


if __name__ == "__main__":
    main()
