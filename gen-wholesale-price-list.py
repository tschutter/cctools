#!/usr/bin/env python

"""
Generates a wholesale price list in spreadsheet form.
"""

import ConfigParser
import cctools
import itertools
import math
import openpyxl  # sudo apt-get install python-openpyxl
import optparse
import os
import sys
import datetime

# Cell style constants.
ALIGNMENT_HORIZONTAL_RIGHT = openpyxl.style.Alignment.HORIZONTAL_RIGHT
ALIGNMENT_VERTICAL_TOP = openpyxl.style.Alignment.VERTICAL_TOP

def set_cell(
    worksheet,
    row,
    col,
    value,
    bold = None,
    alignment_horizontal = None,
    alignment_vertical = None
):
    """Set cell value and style."""
    cell = worksheet.cell(row=row, column=col)
    cell.value = value
    if bold != None:
        cell.style.font.bold = bold
    if alignment_horizontal != None:
        cell.style.alignment.horizontal = alignment_horizontal
    if alignment_vertical != None:
        cell.style.alignment.vertical = alignment_vertical
    return cell


def col_letter(col):
    """Return column letter for given column."""
    return chr(ord("A") + col)


def row_number(row):
    """Return row number for given row."""
    return row + 1


def add_title(worksheet):
    """Add worksheet title."""
    now = datetime.datetime.now()
    cell_text = now.strftime("Wholesale Prices as of %Y-%m-%d")
    style = set_cell(worksheet, 0, 0, cell_text, bold=True).style
    style.font.size = 20
    # merge_cells not supported by openpyxl-1.5.6 (Ubuntu 12.04)
    #worksheet.merge_cells(start_row=0, start_column=0, end_row=0, end_column=2)
    worksheet.row_dimensions[1].height = 25


def add_products(options, worksheet, row, cc_browser, products):
    """Add row for each product."""
    col_category = 0
    col_description = 1
    col_sku = 2
    col_price = 3
    col_qty = 4
    col_size = 5

    # Add header row.
    set_cell(
        worksheet,
        row,
        col_category,
        "Category",
        bold=True
    )
    set_cell(worksheet, row, col_description, "Description", bold=True)
    set_cell(
        worksheet,
        row,
        col_sku,
        "SKU",
        bold=True,
        alignment_horizontal=ALIGNMENT_HORIZONTAL_RIGHT
    )
    set_cell(
        worksheet,
        row,
        col_price,
        "Price",
        bold=True,
        alignment_horizontal=ALIGNMENT_HORIZONTAL_RIGHT
    )
    set_cell(
        worksheet,
        row,
        col_qty,
        "Qty",
        bold=True,
        alignment_horizontal=ALIGNMENT_HORIZONTAL_RIGHT
    )
    set_cell(
        worksheet,
        row,
        col_size,
        "Size",
        bold=True
    )
    row += 1

    # Sort products by category, product_name.
    products = sorted(products, key=cc_browser.sort_key_by_category_and_name)

    # Group products by category.
    for _, product_group in itertools.groupby(
        products,
        key=cc_browser.sort_key_by_category
    ):
        # Add product rows.
        for product in product_group:
            if product["Discontinued Item"] == "Y":
                continue
            description = "%s: %s" % (
                product["Product Name"],
                cctools.html_to_plain_text(product["Teaser"])
            )
            set_cell(worksheet, row, col_category, product["Category"])
            set_cell(worksheet, row, col_description, description)
            set_cell(worksheet, row, col_sku, product["SKU"])
            online_price = product["Price"]
            rounded_price = math.floor(float(online_price) + 0.5)
            wholesale_price = rounded_price * options.wholesale_fraction
            style = set_cell(worksheet, row, col_price, wholesale_price).style
            style.number_format.format_code = "0.00"
            set_cell(worksheet, row, col_size, product["Size"])
            row += 1

    # Set column widths.
    worksheet.column_dimensions[col_letter(col_category)].width = 20
    worksheet.column_dimensions[col_letter(col_description)].width = 65
    worksheet.column_dimensions[col_letter(col_sku)].width = 6
    worksheet.column_dimensions[col_letter(col_price)].width = 5
    worksheet.column_dimensions[col_letter(col_qty)].width = 5
    worksheet.column_dimensions[col_letter(col_size)].width = 8


def add_price_list(options, cc_browser, products, worksheet):
    """Create the Wholesale Prices worksheet."""

    # Prepare worksheet.
    worksheet.title = "Wholesale Prices"

    # Add title.
    add_title(worksheet)

    # Blank row.
    row = 2

    # Add products.
    add_products(
        options,
        worksheet,
        row,
        cc_browser,
        products
    )


def generate_xlsx(options, cc_browser, products):
    """Generate the XLS file."""

    # Construct a document.
    workbook = openpyxl.workbook.Workbook()

    # Create PO-Invoice worksheet.
    add_price_list(
        options,
        cc_browser,
        products,
        workbook.worksheets[0]
    )

    # Write to file.
    workbook.save(options.xlsx_filename)


def main():
    """main"""
    default_config = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "cctools.cfg"
    )
    now = datetime.datetime.now()
    default_xlsx_filename = now.strftime("%Y-%m-%d-WholesalePrices.xlsx")

    option_parser = optparse.OptionParser(
        usage="usage: %prog [options]\n" +
        "  Generates a Purchase Order / Commercial Invoice."
    )
    option_parser.add_option(
        "--config",
        action="store",
        dest="config",
        metavar="FILE",
        default=default_config,
        help="configuration filename (default=%default)"
    )
    option_parser.add_option(
        "--wholesale-fraction",
        metavar="FRAC",
        default=0.5,
        help="wholesale price fraction (default=%default)"
    )
    option_parser.add_option(
        "--outfile",
        action="store",
        dest="xlsx_filename",
        metavar="FILE",
        default=default_xlsx_filename,
        help="output XLSX filename (default=%default)"
    )
    option_parser.add_option(
        "--verbose",
        action="store_true",
        default=False,
        help="display progress messages"
    )

    # Parse command line arguments.
    (options, args) = option_parser.parse_args()
    if len(args) != 0:
        option_parser.error("invalid argument")

    # Read config file.
    config = ConfigParser.RawConfigParser()
    config.readfp(open(options.config))

    # Create a connection to CoreCommerce.
    cc_browser = cctools.CCBrowser(
        config.get("website", "host"),
        config.get("website", "site"),
        config.get("website", "username"),
        config.get("website", "password"),
        verbose=options.verbose
    )

    # Fetch products list.
    products = cc_browser.get_products()

    # Generate spreadsheet.
    if options.verbose:
        sys.stderr.write("Generating %s\n" % options.xlsx_filename)
    generate_xlsx(options, cc_browser, products)

    if options.verbose:
        sys.stderr.write("Generation complete\n")
    return 0


if __name__ == "__main__":
    main()
