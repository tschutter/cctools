#!/usr/bin/env python2

"""
Generate an inventory tracking spreadsheet.
"""

import ConfigParser
import argparse
import datetime
import itertools
import logging
import os

import openpyxl  # sudo pip install openpyxl
import notify_send_handler

import cctools


def set_cell(
    worksheet,
    row,
    col,
    value,
    font_bold=False,
    font_size=11,
    alignment_horizontal="general",
    alignment_vertical="bottom"
):
    """Set cell value and style."""
    cell = worksheet.cell(row=row, column=col)
    cell.value = value
    cell.font = openpyxl.styles.Font(bold=font_bold, size=font_size)
    cell.alignment = openpyxl.styles.Alignment(
        horizontal=alignment_horizontal,
        vertical=alignment_vertical
    )


def col_letter(col):
    """Return column letter given number."""
    return chr(ord("A") + col - 1)


def add_inventory_to_worksheet(args, inventory, worksheet):
    """Add inventory to a worksheet."""

    # Create header row.
    row = 3
    col = 1
    set_cell(worksheet, row, col, "SKU", font_bold=True)
    worksheet.column_dimensions[col_letter(col)].width = 14
    col += 1

    set_cell(worksheet, row, col, "Product", font_bold=True)
    worksheet.column_dimensions[col_letter(col)].width = 40
    col += 1

    set_cell(
        worksheet,
        row,
        col,
        "Current",
        font_bold=True,
        alignment_horizontal="right"
    )
    worksheet.column_dimensions[col_letter(col)].width = 8
    col += 1

    set_cell(
        worksheet,
        row,
        col,
        "Initial",
        font_bold=True,
        alignment_horizontal="right"
    )
    worksheet.column_dimensions[col_letter(col)].width = 6

    set_cell(
        worksheet,
        1,
        col,
        "Outlet",
        font_bold=True,
        alignment_horizontal="right"
    )
    worksheet.column_dimensions[col_letter(col)].width = 8

    set_cell(
        worksheet,
        2,
        col,
        "Date",
        font_bold=True,
        alignment_horizontal="right"
    )
    worksheet.column_dimensions[col_letter(col)].width = 8
    col += 1

    n_events = 10
    for _ in range(n_events):
        worksheet.column_dimensions[col_letter(col)].width = 11
        set_cell(
            worksheet,
            row,
            col,
            "Change",
            font_bold=True,
            alignment_horizontal="right"
        )
        cell = worksheet.cell(row=row, column=col)
        cell.number_format = "yyyy-mm-dd"
        col += 1

    # Create data rows.
    for itemid, (sku, name, level, enabled) in enumerate(inventory):
        row = itemid + 4
        col = 1
        set_cell(worksheet, row, col, sku, alignment_horizontal="left")
        col += 1
        set_cell(worksheet, row, col, name)
        col += 1
        current = "=SUM({0}{1}:{2}{1}".format(
            col_letter(col + 1),
            row,
            col_letter(col + 1 + n_events)
        )
        set_cell(worksheet, row, col, current)
        col += 1
        set_cell(worksheet, row, col, int(level))

    # Freeze the first four columns.
    worksheet.freeze_panes = "E1"


def generate_xlsx(args, inventory):
    """Generate the XLS file."""

    # Construct a document.
    workbook = openpyxl.workbook.Workbook()

    # Create and fill in a tab for each category.
    for index, (category, products) in enumerate(inventory):
        # Get or create worksheet (tab).
        if index == 0:
            worksheet = workbook.worksheets[0]
            worksheet.title = category
        else:
            worksheet = workbook.create_sheet(category)

        # Add the products to the worksheet.
        add_inventory_to_worksheet(args, products, worksheet)

    # Write to file.
    workbook.save(args.xlsx_filename)


def fetch_inventory(args, config):
    """Fetch inventory data from CoreCommerce."""

    # Create a connection to CoreCommerce.
    cc_browser = cctools.CCBrowser(
        config.get("website", "base_url"),
        config.get("website", "username"),
        config.get("website", "password")
    )

    # Get list of products.
    products = cc_browser.get_products()

    # Sort products by category, product_name.
    products = sorted(products, key=cc_browser.product_key_by_cat_and_name)

    # Get list of variants.
    variants = cc_browser.get_variants()
    variants = sorted(variants, key=cc_browser.variant_key)

    # Group products by category.
    inventory = []
    for _, product_group in itertools.groupby(
        products,
        key=cc_browser.product_key_by_category
    ):
        # Assemble product data for the product_group.
        category_products = None
        for product in product_group:
            if product["Available"] == "N":
                continue

            category_name = product["Category"]
            if category_products is None:
                category_products = []
                inventory.append((category_name, category_products))

            product_sku = product["SKU"]
            product_name = product["Product Name"]
            product_level = product["Inventory Level"]
            if product["Track Inventory"] == "By Product":
                enabled = product["Available"]
                category_products.append(
                    (product_sku, product_name, product_level, enabled)
                )
            else:
                for variant in variants:
                    if product_sku == variant["Product SKU"]:
                        variant_sku = variant["Variant SKU"]
                        if variant_sku == "":
                            sku = product_sku
                        else:
                            sku = "{}-{}".format(product_sku, variant_sku)
                        answer = variant["Variant Name"]
                        if answer == "Assorted":
                            continue
                        name = "{} ({})".format(product_name, answer)
                        variant_inventory_level = variant[
                            "Variant Inventory Level"
                        ]
                        enabled = variant["Variant Enabled"]
                        category_products.append(
                            (
                                sku,
                                name,
                                variant_inventory_level,
                                enabled
                            )
                        )

    return inventory


def main():
    """main"""
    default_config = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "cctools.cfg"
    )
    now = datetime.datetime.now()
    default_xlsx_filename = now.strftime("%Y-%m-%d-InventoryTracker.xlsx")

    arg_parser = argparse.ArgumentParser(
        description="Generates an inventory report."
    )
    arg_parser.add_argument(
        "--config",
        metavar="FILE",
        default=default_config,
        help="configuration filename (default=%(default)s)"
    )
    arg_parser.add_argument(
        "--outfile",
        dest="xlsx_filename",
        metavar="FILE",
        default=default_xlsx_filename,
        help="output XLSX filename (default=%(default)s)"
    )
    arg_parser.add_argument(
        "--sort",
        dest="sort",
        choices=["SKU", "CAT/PROD"],
        default="SKU",
        help="sort order (default=%(default)s)"
    )
    arg_parser.add_argument(
        "--verbose",
        action="store_true",
        default=False,
        help="display progress messages"
    )

    # Parse command line arguments.
    args = arg_parser.parse_args()

    # Configure logging.
    logging.basicConfig(
        level=logging.INFO if args.verbose else logging.WARNING
    )
    logger = logging.getLogger()

    # Also log using notify-send if it is available.
    if notify_send_handler.NotifySendHandler.is_available():
        logger.addHandler(
            notify_send_handler.NotifySendHandler(
                os.path.splitext(os.path.basename(__file__))[0]
            )
        )

    # Read config file.
    config = ConfigParser.RawConfigParser()
    config.readfp(open(args.config))

    # Get inventory info.
    inventory = fetch_inventory(args, config)

    # Create spreadsheet.
    logger.debug("Generating %s", args.xlsx_filename)
    generate_xlsx(args, inventory)

    logger.debug("Generation complete")
    return 0

if __name__ == "__main__":
    main()
